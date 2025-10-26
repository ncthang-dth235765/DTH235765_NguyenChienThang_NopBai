#Câu 8: Xử lý đọc Excel File
from openpyxl import load_workbook

# Mở file Excel có sẵn
wb = load_workbook('demo.xlsx')

# In ra danh sách tên các sheet
print(wb.sheetnames)

# Lấy sheet đầu tiên
ws = wb[wb.sheetnames[0]]

# Duyệt từng dòng và in các giá trị trong mỗi ô
for row in ws.values:
    for value in row:
        print(value, "\t", end='')
    print()
