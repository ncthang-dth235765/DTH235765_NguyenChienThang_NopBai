#Câu 11: Xử lý Excel File - Viết phần mềm Quản Lý Nhân Viên
from openpyxl import Workbook, load_workbook

# ==========================
# 1️⃣ HÀM LƯU FILE EXCEL
# ==========================
def LuuNhanVienExcel(filename, ds_nv):
    wb = Workbook()
    ws = wb.active
    ws.title = "NhanVien"

    # Tạo tiêu đề cột
    ws.append(["STT", "Mã NV", "Tên", "Tuổi"])

    # Ghi dữ liệu
    for i, nv in enumerate(ds_nv, start=1):
        ws.append([i, nv["ma"], nv["ten"], nv["tuoi"]])

    wb.save(filename)
    print("✅ Đã lưu danh sách nhân viên vào file:", filename)


# ==========================
# 2️⃣ HÀM ĐỌC FILE EXCEL
# ==========================
def DocNhanVienExcel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    ds_nv = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # bỏ dòng tiêu đề
        stt, ma, ten, tuoi = row
        ds_nv.append({"ma": ma, "ten": ten, "tuoi": tuoi})

    return ds_nv


# ==========================
# 3️⃣ HÀM SẮP XẾP TUỔI
# ==========================
def SapXepTheoTuoi(ds_nv):
    ds_nv.sort(key=lambda nv: nv["tuoi"])
    print("✅ Đã sắp xếp nhân viên theo tuổi tăng dần.")


# ==========================
# 4️⃣ HÀM HIỂN THỊ DANH SÁCH
# ==========================
def XuatNhanVien(ds_nv):
    print("\n{:<10}{:<20}{:<10}".format("Mã NV", "Tên", "Tuổi"))
    print("-" * 40)
    for nv in ds_nv:
        print("{:<10}{:<20}{:<10}".format(nv["ma"], nv["ten"], nv["tuoi"]))
    print()


# ==========================
# 5️⃣ CHƯƠNG TRÌNH CHÍNH
# ==========================
def main():
    ds_nv = [
        {"ma": "NV1", "ten": "An", "tuoi": 20},
        {"ma": "NV2", "ten": "Lành", "tuoi": 19},
        {"ma": "NV3", "ten": "Thảo", "tuoi": 21},
        {"ma": "NV4", "ten": "Thọ", "tuoi": 18},
        {"ma": "NV5", "ten": "Phúc", "tuoi": 24}
    ]

    filename = "nhanvien.xlsx"

    while True:
        print("\n=== QUẢN LÝ NHÂN VIÊN ===")
        print("1. Lưu danh sách nhân viên vào file Excel")
        print("2. Đọc danh sách nhân viên từ file Excel")
        print("3. Sắp xếp nhân viên theo tuổi tăng dần")
        print("4. Xuất danh sách nhân viên")
        print("5. Thoát")
        chon = input("👉 Chọn chức năng: ")

        if chon == "1":
            LuuNhanVienExcel(filename, ds_nv)
        elif chon == "2":
            ds_nv = DocNhanVienExcel(filename)
            print("✅ Đã đọc danh sách từ file Excel!")
        elif chon == "3":
            SapXepTheoTuoi(ds_nv)
        elif chon == "4":
            XuatNhanVien(ds_nv)
        elif chon == "5":
            print("👋 Kết thúc chương trình!")
            break
        else:
            print("❌ Lựa chọn không hợp lệ!")


if __name__ == "__main__":
    main()
